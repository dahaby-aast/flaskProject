import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import Tk, Toplevel, filedialog, messagebox, simpledialog, Listbox
import sys
import os
import pandas as pd
import warnings
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore", category=SyntaxWarning)

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller."""
    base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
    return os.path.join(base_path, relative_path)


class DashboardApp:
    def __init__(self, root):
        self.root = root
        self.style = ttk.Style()  # Create a ttkbootstrap Style object
        self.current_theme = "darkly"
        self.style.theme_use(self.current_theme)  # Apply the initial theme
        self.create_custom_styles()  # Apply custom styles
        self.root.title("Dashboard")
        self.root.geometry("400x720")
        self.root.resizable(True, True)
        self.create_widgets()

    def create_custom_styles(self):
        # Custom styles to apply to all themes
        self.style.configure("TButton", font=("Helvetica", 12, "bold"))
        self.style.configure("TLabel", foreground="#fdfdfd")

    def create_widgets(self):
        self.logo = ttk.PhotoImage(file=resource_path("uploads/logo1.png"))
        ttk.Label(self.root, image=self.logo).pack(pady=10)
        ttk.Label(self.root, text="My Dashboard", font=("Helvetica", 14, "bold"), bootstyle="primary").pack(pady=20)

        current_date = datetime.now().strftime("%d-%m-%Y")

        # Date and Time Frames
        date_frame = ttk.Frame(self.root, bootstyle="info")
        date_frame.pack(pady=10, padx=20, fill="x")
        self.date_label = ttk.Label(date_frame, text=f"Date: {current_date}",
                                    font=("Helvetica", 12, "bold"), bootstyle="inverse-info")
        self.date_label.pack(pady=5)

        time_frame = ttk.Frame(self.root, bootstyle="info")
        time_frame.pack(pady=10, padx=20, fill="x")
        self.time_label = ttk.Label(time_frame, text="", font=("Helvetica", 12, "bold"), bootstyle="inverse-info")
        self.time_label.pack(pady=5)

        self.update_time()  # Start updating time

        ttk.Button(self.root, text="Open Add Column App", command=self.open_add_column_app, bootstyle=SUCCESS).pack(pady=10, padx=20, fill="x")
        ttk.Button(self.root, text="Open File Combiner App", command=self.open_file_combiner_app, bootstyle=INFO).pack(pady=10, padx=20, fill="x")
        ttk.Button(self.root, text="Open File", command=self.open_file, bootstyle=PRIMARY).pack(pady=10, padx=20, fill="x")
        ttk.Button(self.root, text="Toggle Theme", command=self.toggle_theme, bootstyle=WARNING).pack(pady=10, padx=20, fill="x")
        ttk.Button(self.root, text="Exit", command=self.root.quit, bootstyle=DANGER).pack(pady=10, padx=20, fill="x")

        ttk.Label(self.root, text="Все права защищены @ Дахаби 2024", font=("Helvetica", 12, "bold"),
                  bootstyle="secondary", anchor="w", padding=(10, 5), foreground="#8D9FB1").pack(side="bottom", fill="x")

    def update_time(self):
        self.time_label.config(text=f"Time: {datetime.now().strftime('%H:%M:%S')}")
        self.root.after(1000, self.update_time)

    def open_add_column_app(self):
        self._open_app_window("Add Column App", AddColumnApp)

    def open_file_combiner_app(self):
        self._open_app_window("File Combiner App", FileCombinerApp)

    def open_file(self):
        file_path = filedialog.askopenfilename(title="Select a File", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if file_path:
            try:
                os.startfile(file_path)
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {e}")

    def _open_app_window(self, title, app_class):
        self.root.iconify()
        top = Toplevel(self.root)
        top.title(title)
        top.geometry("800x600")
        app_class(top)
        top.protocol("WM_DELETE_WINDOW", lambda: self.on_close_app(top))

    def on_close_app(self, window):
        window.destroy()
        self.root.deiconify()

    def toggle_theme(self):
        available_themes = ["sandstone", "flatly", "darkly", "cosmo", "minty", "superhero"]
        index = available_themes.index(self.current_theme)
        self.current_theme = available_themes[(index + 1) % len(available_themes)]
        self.style.theme_use(self.current_theme)
        self.create_custom_styles()


class AddColumnApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Add Column to Multiple Excel Files")
        self.root.geometry("900x600")
        self.root.resizable(True, True)
        self.file_paths = [None] * 6
        self.text_entries = [ttk.StringVar() for _ in range(6)]
        self.records_labels = [None] * 6
        self.create_widgets()

    def create_widgets(self):
        ttk.Label(self.root, text="Add 'Location' Column to Excel Files", font=("Helvetica", 12, "bold"), bootstyle="primary").grid(row=0, column=0, columnspan=3, pady=20)

        for i in range(6):
            frame = ttk.Frame(self.root, borderwidth=2, relief="flat")
            frame.grid(row=i + 1, column=0, columnspan=3, pady=5, padx=10, sticky="ew")

            file_frame = ttk.Frame(frame)
            file_frame.grid(row=0, column=0, padx=(0, 10), sticky="w")
            ttk.Button(file_frame, text=f"Upload File {i + 1}", command=lambda i=i: self.upload_file(i), bootstyle=INFO).pack(side=LEFT)
            label = ttk.Label(file_frame, text="No file selected", bootstyle=SECONDARY, anchor="w", font=("Helvetica", 12, "bold"), padding=(10, 0, 0, 0))
            label.pack(side=LEFT, fill="x", expand=True)
            setattr(self, f'file{i + 1}_label', label)

            self.records_labels[i] = ttk.Label(file_frame, text="Records: N/A", font=("Helvetica", 12, "bold"), bootstyle=SECONDARY, anchor="w")
            self.records_labels[i].pack(side=LEFT, padx=(10, 0))

            column_frame = ttk.Frame(frame)
            column_frame.grid(row=0, column=1, padx=(10, 0), sticky="e")
            ttk.Label(column_frame, text=f"Location {i + 1}:").pack(anchor="w")
            options = ['Maritime_lab 232', 'CS_lab 206', 'CS_lab 404']
            combo_box = ttk.Combobox(column_frame, values=options, textvariable=self.text_entries[i])
            combo_box.pack(fill="x")

        ttk.Button(self.root, text="Add Column to Files", command=self.add_column_to_files, bootstyle=SUCCESS).grid(row=7, column=0, columnspan=3, pady=15, padx=20)
        button_frame = ttk.Frame(self.root)
        button_frame.grid(row=8, column=0, columnspan=3, pady=10, padx=20, sticky="ew")
        ttk.Button(button_frame, text="Clear", command=self.clear, bootstyle=WARNING).pack(side=LEFT, fill="x", expand=True, padx=5)
        ttk.Button(button_frame, text="Exit", command=self.root.destroy, bootstyle=DANGER).pack(side=LEFT, fill="x", expand=True, padx=5)

    def upload_file(self, index):
        files = filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx *.xls")])
        if files:
            for i in range(min(len(files), 6)):
                self.file_paths[i] = files[i]
                getattr(self, f'file{i + 1}_label').config(text=os.path.basename(files[i]))
                self.update_record_count(i)
            for i in range(len(files), 6):
                self.file_paths[i] = None
                getattr(self, f'file{i + 1}_label').config(text="No file selected")
                self.records_labels[i].config(text="Records: N/A")

    def update_record_count(self, index):
        try:
            df = pd.read_excel(self.file_paths[index])
            self.records_labels[index].config(text=f"Records: {len(df)}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while reading file {index + 1}: {e}")

    def add_column_to_files(self):
        if not any(self.file_paths):
            messagebox.showwarning("Warning", "No files selected.")
            return

        # Ask for the prefix only once
        prefix = simpledialog.askstring("Input",
                                        "Enter the prefix for the saved files:")
        if not prefix:
            messagebox.showwarning("Warning", "No prefix provided.")
            return

        try:
            for i, file_path in enumerate(self.file_paths):
                if file_path:
                    df = pd.read_excel(file_path)
                    df['Location'] = self.text_entries[i].get()

                    # Automatically save the file using the prefix and the original filename
                    original_file_name = os.path.basename(file_path)
                    new_file_name = f"{prefix}_{original_file_name}"
                    new_file_path = os.path.join(os.path.dirname(file_path),
                                                 new_file_name)
                    df.to_excel(new_file_path, index=False)

            # Show a single success message after all files are saved
            messagebox.showinfo("Success",
                                "All files have been saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def clear(self):
        for i in range(6):
            self.file_paths[i] = None
            getattr(self, f'file{i + 1}_label').config(text="No file selected")
            self.records_labels[i].config(text="Records: N/A")
            self.text_entries[i].set("")


class FileCombinerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Combine Excel Files")
        self.root.geometry("600x400")
        self.root.resizable(True, True)
        self.file_paths = []
        self.create_widgets()

    def create_widgets(self):
        ttk.Label(self.root, text="Combine Excel Files", font=("Helvetica", 12, "bold"), bootstyle="primary").pack(pady=20)

        self.listbox = Listbox(self.root, selectmode="multiple", height=10, width=80)
        self.listbox.pack(pady=10, padx=20)

        ttk.Button(self.root, text="Add Files", command=self.add_files, bootstyle=INFO).pack(pady=10, padx=20)
        ttk.Button(self.root, text="Combine Files", command=self.combine_files, bootstyle=SUCCESS).pack(pady=10, padx=20)
        ttk.Button(self.root, text="Clear", command=self.clear_files, bootstyle=WARNING).pack(pady=10, padx=20)
        ttk.Button(self.root, text="Exit", command=self.root.destroy, bootstyle=DANGER).pack(pady=10, padx=20)

    def add_files(self):
        files = filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx *.xls")])
        if files:
            for file in files:
                if file not in self.file_paths:
                    self.file_paths.append(file)
                    self.listbox.insert("end", os.path.basename(file))

    def combine_files(self):
        if not self.file_paths:
            messagebox.showwarning("Warning", "No files selected.")
            return

        combined_df = pd.concat((pd.read_excel(file) for file in self.file_paths), ignore_index=True)

        # Remove 'Email' and 'Date' columns if they exist
        for column in ['Email', 'Date']:
            if column in combined_df.columns:
                combined_df.drop(columns=[column], inplace=True)

        # Sort by 'RegNum' column if it exists
        if 'RegNum' in combined_df.columns:
            combined_df.sort_values(by='RegNum', inplace=True)

        # Ask for the file name and save the file
        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_file:
            try:
                self.save_combined_file(output_file, combined_df)
                messagebox.showinfo("Success", f"Files combined and saved to {output_file}")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while saving the combined file: {e}")

    def save_combined_file(self, output_file, combined_df):
        wb = Workbook()
        ws = wb.active
        ws.title = "Combined Data"

        # Append the combined dataframe to the worksheet
        for r in dataframe_to_rows(combined_df, index=False, header=True):
            ws.append(r)

        # Determine the range of cells with data
        max_row = ws.max_row
        max_col = ws.max_column

        # Define the table range dynamically
        table_range = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"

        # Define the table in the worksheet with the dynamic range
        table = Table(displayName="CombinedTable", ref=table_range)

        # Define the style for the table
        style = TableStyleInfo(
            name="TableStyleMedium18",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        # Add the table to the worksheet
        ws.add_table(table)

        # Save the workbook
        wb.save(output_file)

    def clear_files(self):
        self.file_paths = []
        self.listbox.delete(0, "end")



if __name__ == "__main__":
    root = Tk()
    app = DashboardApp(root)
    root.mainloop()
